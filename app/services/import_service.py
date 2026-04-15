"""Import Service - Data import from Excel, CSV and other formats"""

import os
from datetime import datetime
from typing import Optional

from werkzeug.utils import secure_filename

from app.models import Lancamento, Obra, db
from app.utils.dates import parse_date


class ImportService:
    """Service for importing data from files"""

    ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}

    @staticmethod
    def allowed_file(filename: str) -> bool:
        """Check if file extension is allowed"""
        return (
            '.' in filename
            and filename.rsplit('.', 1)[1].lower() in ImportService.ALLOWED_EXTENSIONS
        )

    @staticmethod
    def importar_lancamentos(
        empresa_id: int, file_path: str, dry_run: bool = False
    ) -> tuple[int, int, list[str]]:
        """
        Import lancamentos from Excel/CSV file.

        Args:
            empresa_id: Empresa ID
            file_path: Path to the file
            dry_run: If True, only validate without saving

        Returns:
            Tuple (imported_count, error_count, errors)
        """
        import openpyxl

        errors = []
        imported = 0
        error_count = 0

        try:
            # Load workbook
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active

            # Skip header row
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return 0, 0, ['Arquivo vazio']

            header = rows[0]
            data_rows = rows[1:]

            for row_num, row in enumerate(data_rows, start=2):
                try:
                    # Parse row data (adjust column indices based on your template)
                    dados = ImportService._parse_lancamento_row(row, header)

                    if not dados:
                        error_count += 1
                        errors.append(f'Linha {row_num}: Dados inválidos')
                        continue

                    # Validate required fields
                    if not dados.get('descricao'):
                        error_count += 1
                        errors.append(f'Linha {row_num}: Descrição obrigatória')
                        continue

                    if not dados.get('valor') or dados['valor'] <= 0:
                        error_count += 1
                        errors.append(f'Linha {row_num}: Valor deve ser positivo')
                        continue

                    if dry_run:
                        imported += 1
                        continue

                    # Create lancamento
                    lancamento = Lancamento(
                        empresa_id=empresa_id,
                        obra_id=dados.get('obra_id'),
                        descricao=dados['descricao'],
                        categoria=dados.get('categoria'),
                        tipo=dados.get('tipo', 'Despesa'),
                        valor=dados['valor'],
                        data=dados.get('data'),
                        forma_pagamento=dados.get('forma_pagamento'),
                        status_pagamento=dados.get('status_pagamento', 'Pago'),
                        observacoes=dados.get('observacoes'),
                        documento=dados.get('documento'),
                    )
                    db.session.add(lancamento)
                    imported += 1

                except Exception as e:
                    error_count += 1
                    errors.append(f'Linha {row_num}: {e!s}')

            if not dry_run:
                db.session.commit()

            wb.close()

        except Exception as e:
            errors.append(f'Erro ao processar arquivo: {e!s}')
            if not dry_run:
                db.session.rollback()

        return imported, error_count, errors

    @staticmethod
    def _parse_lancamento_row(row, header) -> dict | None:
        """Parse a row from Excel/CSV into lancamento data dict"""
        try:
            # Map header names to field names
            field_mapping = {
                'descricao': ['descricao', 'descrição', 'descrição', 'description'],
                'valor': ['valor', 'value', 'amount', 'valor (R$)'],
                'data': ['data', 'date', 'data_vencimento', 'vencimento'],
                'categoria': ['categoria', 'category', 'tipo_despesa'],
                'tipo': ['tipo', 'type', 'tipo_lancamento'],
                'obra': ['obra', 'projeto', 'work', 'nome_obra'],
                'forma_pagamento': ['forma_pagamento', 'pagamento', 'payment_method'],
                'status_pagamento': ['status', 'status_pagamento', 'payment_status'],
                'observacoes': ['observacoes', 'obs', 'notas', 'notes'],
                'documento': ['documento', 'doc', 'numero_documento'],
            }

            dados = {}

            # Create reverse mapping
            col_index = {}
            for idx, col_name in enumerate(header):
                if col_name:
                    col_index[str(col_name).lower().strip()] = idx

            # Map fields
            for field, possible_names in field_mapping.items():
                for name in possible_names:
                    if name.lower() in col_index:
                        value = row[col_index[name.lower()]]
                        if value is not None and str(value).strip():
                            dados[field] = str(value).strip()
                        break

            # Special handling for numeric fields
            if 'valor' in dados:
                try:
                    dados['valor'] = float(str(dados['valor']).replace(',', '.'))
                except (ValueError, TypeError):
                    dados['valor'] = 0

            # Special handling for date
            if 'data' in dados:
                dados['data'] = parse_date(dados['data'])

            # Handle obra_id lookup
            if 'obra' in dados:
                obra = Obra.query.filter_by(nome=dados['obra']).first()
                if obra:
                    dados['obra_id'] = obra.id
                del dados['obra']

            return dados if dados else None

        except Exception:
            return None

    @staticmethod
    def importar_obras(
        empresa_id: int, file_path: str, dry_run: bool = False
    ) -> tuple[int, int, list[str]]:
        """
        Import obras from Excel/CSV file.

        Args:
            empresa_id: Empresa ID
            file_path: Path to the file
            dry_run: If True, only validate without saving

        Returns:
            Tuple (imported_count, error_count, errors)
        """
        import openpyxl

        errors = []
        imported = 0
        error_count = 0

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return 0, 0, ['Arquivo vazio']

            header = rows[0]
            data_rows = rows[1:]

            for row_num, row in enumerate(data_rows, start=2):
                try:
                    dados = ImportService._parse_obra_row(row, header)

                    if not dados:
                        error_count += 1
                        errors.append(f'Linha {row_num}: Dados inválidos')
                        continue

                    # Validate required fields
                    if not dados.get('nome'):
                        error_count += 1
                        errors.append(f'Linha {row_num}: Nome obrigatório')
                        continue

                    if dry_run:
                        imported += 1
                        continue

                    # Create obra
                    obra = Obra(
                        empresa_id=empresa_id,
                        nome=dados['nome'],
                        descricao=dados.get('descricao'),
                        endereco=dados.get('endereco'),
                        orcamento_previsto=dados.get('orcamento_previsto', 0),
                        data_inicio=dados.get('data_inicio'),
                        data_fim_prevista=dados.get('data_fim_prevista'),
                        status=dados.get('status', 'Planejamento'),
                        progresso=dados.get('progresso', 0),
                        responsavel=dados.get('responsavel'),
                        cliente=dados.get('cliente'),
                    )
                    db.session.add(obra)
                    imported += 1

                except Exception as e:
                    error_count += 1
                    errors.append(f'Linha {row_num}: {e!s}')

            if not dry_run:
                db.session.commit()

            wb.close()

        except Exception as e:
            errors.append(f'Erro ao processar arquivo: {e!s}')
            if not dry_run:
                db.session.rollback()

        return imported, error_count, errors

    @staticmethod
    def _parse_obra_row(row, header) -> dict | None:
        """Parse a row from Excel/CSV into obra data dict"""
        try:
            field_mapping = {
                'nome': ['nome', 'obra', 'project', 'nome_obra'],
                'descricao': ['descricao', 'descrição', 'description'],
                'endereco': ['endereco', 'endereço', 'address', 'local'],
                'orcamento_previsto': ['orcamento', 'orçamento', 'budget', 'valor'],
                'data_inicio': ['data_inicio', 'inicio', 'start_date'],
                'data_fim_prevista': ['data_fim', 'fim', 'end_date', 'previsao_fim'],
                'status': ['status', 'situacao', 'situação'],
                'progresso': ['progresso', 'progress', '%'],
                'responsavel': ['responsavel', 'responsável', 'manager'],
                'cliente': ['cliente', 'client', 'contratante'],
            }

            dados = {}
            col_index = {}
            for idx, col_name in enumerate(header):
                if col_name:
                    col_index[str(col_name).lower().strip()] = idx

            for field, possible_names in field_mapping.items():
                for name in possible_names:
                    if name.lower() in col_index:
                        value = row[col_index[name.lower()]]
                        if value is not None and str(value).strip():
                            dados[field] = str(value).strip()
                        break

            # Numeric fields
            for field in ['orcamento_previsto', 'progresso']:
                if field in dados:
                    try:
                        dados[field] = float(str(dados[field]).replace(',', '.'))
                    except (ValueError, TypeError):
                        dados[field] = 0

            # Date fields
            for field in ['data_inicio', 'data_fim_prevista']:
                if field in dados:
                    dados[field] = parse_date(dados[field])

            return dados if dados else None

        except Exception:
            return None

    @staticmethod
    def save_upload_file(file, upload_folder: str) -> tuple[str, str]:
        """
        Save uploaded file securely.

        Args:
            file: File object from request.files
            upload_folder: Target folder

        Returns:
            Tuple (saved_path, error_message)
        """
        if not file or not file.filename:
            return '', 'Nenhum arquivo selecionado'

        if not ImportService.allowed_file(file.filename):
            return '', f'Formato não suportado. Use: {", ".join(ImportService.ALLOWED_EXTENSIONS)}'

        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'{timestamp}_{filename}'

        os.makedirs(upload_folder, exist_ok=True)
        file_path = os.path.join(upload_folder, filename)
        file.save(file_path)

        return file_path, ''

    @staticmethod
    def get_import_template_path(tipo: str) -> str:
        """Get path to import template file"""
        templates = {
            'lancamentos': 'templates/import/lancamentos_template.xlsx',
            'obras': 'templates/import/obras_template.xlsx',
        }
        return templates.get(tipo, '')
