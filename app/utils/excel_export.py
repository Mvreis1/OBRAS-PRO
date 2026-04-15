"""
Helper para exportacao de dados para Excel (.xlsx)
Usa openpyxl para gerar arquivos Excel reais com formatacao profissional
"""

from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class EstiloExcel:
    """Centraliza todos os estilos de formatacao do Excel"""

    # Cores do tema
    CORES = {
        'header': '6366F1',
        'header_text': 'FFFFFF',
        'subheader': 'E8EAF6',
        'subheader_text': '3F3F46',
        'positive': '22C55E',
        'negative': 'EF4444',
        'neutral': 'F59E0B',
        'alt_row': 'F8FAFC',
        'border': 'E2E8F0',
    }

    # Formatos de numero
    FORMATOS = {
        'currency': '#,##0.00',
        'number': '#,##0',
        'date': 'DD/MM/YYYY',
        'percent': '0.00%',
    }

    @classmethod
    def header_font(cls):
        return Font(name='Calibri', size=11, bold=True, color=f'FF{cls.CORES["header_text"]}')

    @classmethod
    def header_fill(cls):
        return PatternFill(
            start_color=f'FF{cls.CORES["header"]}',
            end_color=f'FF{cls.CORES["header"]}',
            fill_type='solid',
        )

    @classmethod
    def header_align(cls):
        return Alignment(horizontal='center', vertical='center', wrap_text=True)

    @classmethod
    def subheader_font(cls):
        return Font(name='Calibri', size=10, bold=True, color=f'FF{cls.CORES["subheader_text"]}')

    @classmethod
    def subheader_fill(cls):
        return PatternFill(
            start_color=f'FF{cls.CORES["subheader"]}',
            end_color=f'FF{cls.CORES["subheader"]}',
            fill_type='solid',
        )

    @classmethod
    def data_font(cls, bold=False, color=None):
        cor = color or 'FF1E293B'
        return Font(name='Calibri', size=10, bold=bold, color=cor)

    @classmethod
    def alt_fill(cls):
        return PatternFill(
            start_color=f'FF{cls.CORES["alt_row"]}',
            end_color=f'FF{cls.CORES["alt_row"]}',
            fill_type='solid',
        )

    @classmethod
    def thin_border(cls):
        return Border(
            left=Side(style='thin', color=f'FF{cls.CORES["border"]}'),
            right=Side(style='thin', color=f'FF{cls.CORES["border"]}'),
            top=Side(style='thin', color=f'FF{cls.CORES["border"]}'),
            bottom=Side(style='thin', color=f'FF{cls.CORES["border"]}'),
        )

    @classmethod
    def data_align(cls):
        return Alignment(vertical='center', wrap_text=True)


class ExcelExport:
    """
    Classe para exportacao profissional de Excel.

    Uso:
        exporter = ExcelExport()
        exporter.add_sheet('Dados', cabecalhos, dados)
        return exporter.to_response('relatorio.xlsx')
    """

    def __init__(self):
        self.wb = Workbook()
        # Remover sheet padrao
        default_sheet = self.wb.active
        self.wb.remove(default_sheet)

    def add_sheet(self, nome, cabecalhos, dados):
        """
        Adiciona uma planilha ao arquivo.

        Args:
            nome: Nome da planilha (max 31 chars)
            cabecalhos: Lista de tuplas (titulo, tipo) onde tipo pode ser:
                'text', 'number', 'currency', 'date', 'percent'
            dados: Lista de listas com os dados de cada linha
        """
        ws = self.wb.create_sheet(title=nome[:31])

        estilo = EstiloExcel()
        border = estilo.thin_border()
        alt_fill = estilo.alt_fill()

        # Cabecalho
        for col_idx, (titulo, _) in enumerate(cabecalhos, 1):
            cell = ws.cell(row=1, column=col_idx, value=titulo)
            cell.font = estilo.header_font()
            cell.fill = estilo.header_fill()
            cell.alignment = estilo.header_align()
            cell.border = border

        # Dados
        for row_idx, row_data in enumerate(dados, 2):
            for col_idx, valor in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = estilo.data_align()

                # Tipo da coluna
                tipo = cabecalhos[col_idx - 1][1] if col_idx <= len(cabecalhos) else 'text'

                # Aplicar valor baseado no tipo
                if valor is None or valor in {'', '-'}:
                    cell.value = ''
                    cell.font = estilo.data_font()
                elif tipo == 'currency' and isinstance(valor, (int, float)):
                    cell.value = valor
                    cell.number_format = estilo.FORMATOS['currency']
                    # Cor condicional
                    if valor < 0:
                        cell.font = estilo.data_font(
                            bold=True, color=f'FF{estilo.CORES["negative"]}'
                        )
                    elif valor > 0:
                        cell.font = estilo.data_font(
                            bold=True, color=f'FF{estilo.CORES["positive"]}'
                        )
                    else:
                        cell.font = estilo.data_font()
                elif tipo == 'number' and isinstance(valor, (int, float)):
                    cell.value = valor
                    cell.number_format = estilo.FORMATOS['number']
                    cell.font = estilo.data_font()
                elif tipo == 'percent' and isinstance(valor, (int, float)):
                    cell.value = valor / 100.0 if valor > 1 else valor
                    cell.number_format = estilo.FORMATOS['percent']
                    cell.font = estilo.data_font()
                elif tipo == 'date' and isinstance(valor, (date, datetime)):
                    cell.value = valor
                    cell.number_format = estilo.FORMATOS['date']
                    cell.font = estilo.data_font()
                else:
                    cell.value = str(valor) if valor is not None else ''
                    cell.font = estilo.data_font()

                # Linha alternada
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # Auto-width das colunas
        self._ajustar_largura_colunas(ws, cabecalhos)

        # Congelar primeira linha
        ws.freeze_panes = 'A2'

        # Auto-filtro
        ws.auto_filter.ref = f'A1:{get_column_letter(len(cabecalhos))}{len(dados) + 1}'

    def add_summary(self, sheet_nome, resumo):
        """Adiciona linhas de resumo ao final da planilha"""
        ws = self.wb[sheet_nome]
        last_row = ws.max_row + 2  # Pula uma linha
        estilo = EstiloExcel()
        border = estilo.thin_border()

        for i, (label, valor) in enumerate(resumo.items()):
            row = last_row + i

            cell_label = ws.cell(row=row, column=1, value=label)
            cell_label.font = estilo.subheader_font()
            cell_label.fill = estilo.subheader_fill()
            cell_label.border = border

            cell_valor = ws.cell(row=row, column=2, value=valor)
            cell_valor.font = estilo.data_font(bold=True)
            cell_valor.border = border

            # Se for numero, formatar como moeda
            if isinstance(valor, (int, float)):
                cell_valor.number_format = EstiloExcel.FORMATOS['currency']

    def to_bytes(self):
        """Retorna bytes do arquivo Excel"""
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def to_response(self, filename):
        """Retorna response Flask para download"""
        from flask import make_response

        return make_response(
            self.to_bytes(),
            headers={
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': f'attachment; filename={filename}',
            },
        )

    def _ajustar_largura_colunas(self, ws, cabecalhos):
        """Ajusta automaticamente a largura das colunas"""
        for col_idx in range(1, len(cabecalhos) + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            # Largura minima 10, maxima 40
            adjusted_width = min(max(max_length + 2, 10), 40)
            ws.column_dimensions[col_letter].width = adjusted_width


def format_currency_br(value):
    """Formata valor para moeda brasileira"""
    if value is None:
        return ''
    return f'R$ {value:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')


def format_date_br(value):
    """Formata data para padrao brasileiro"""
    if value is None:
        return '-'
    if hasattr(value, 'strftime'):
        return value.strftime('%d/%m/%Y')
    return str(value)


def formatar_moeda():
    return None
