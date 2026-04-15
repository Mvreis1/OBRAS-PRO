"""
Gerador de planilha de exemplo com 20 obras para importação
"""

import random
from datetime import date, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def gerar_planilha_obras_exemplo():
    """
    Gera uma planilha Excel de exemplo com 20 obras fictícias
    para demonstração do sistema OBRAS PRO
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Obras Exemplo'

    # Estilos
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    # Cabeçalhos
    cabecalhos = [
        'Nome da Obra',
        'Cliente',
        'Endereço',
        'Status',
        'Orçamento Previsto (R$)',
        'Data Início',
        'Data Fim Prevista',
        'Progresso (%)',
        'Responsável',
        'Descrição',
    ]

    for col_idx, titulo in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Dados de exemplo - 20 obras
    obras = [
        {
            'nome': 'Residencial Villa Nova',
            'cliente': 'João Silva',
            'endereco': 'Rua das Flores, 123 - São Paulo, SP',
            'status': 'Em Execução',
            'orcamento': 850000.00,
            'dias_inicio': -180,
            'dias_fim': 90,
            'progresso': 65,
            'responsavel': 'Eng. Carlos Mendes',
            'descricao': 'Construção de condomínio residencial com 12 unidades',
        },
        {
            'nome': 'Centro Comercial Plaza',
            'cliente': 'Imobiliária ABC Ltda',
            'endereco': 'Av. Principal, 500 - Centro, Rio de Janeiro, RJ',
            'status': 'Planejamento',
            'orcamento': 2500000.00,
            'dias_inicio': 30,
            'dias_fim': 365,
            'progresso': 10,
            'responsavel': 'Eng. Maria Santos',
            'descricao': 'Centro comercial com 20 lojas e estacionamento',
        },
        {
            'nome': 'Escola Municipal Prof. Lima',
            'cliente': 'Prefeitura Municipal',
            'endereco': 'Rua da Educação, 45 - Belo Horizonte, MG',
            'status': 'Em Execução',
            'orcamento': 1200000.00,
            'dias_inicio': -120,
            'dias_fim': 180,
            'progresso': 40,
            'responsavel': 'Eng. Pedro Oliveira',
            'descricao': 'Reforma e ampliação da escola municipal',
        },
        {
            'nome': 'Hospital Regional Norte',
            'cliente': 'Secretaria de Saúde',
            'endereco': 'Av. da Saúde, 1000 - Porto Alegre, RS',
            'status': 'Planejamento',
            'orcamento': 5000000.00,
            'dias_inicio': 60,
            'dias_fim': 730,
            'progresso': 5,
            'responsavel': 'Eng. Ana Costa',
            'descricao': 'Construção de hospital regional com 100 leitos',
        },
        {
            'nome': 'Galpão Industrial Logística',
            'cliente': 'Transportadora Rápida S/A',
            'endereco': 'Rodovia BR-101, Km 50 - Curitiba, PR',
            'status': 'Concluída',
            'orcamento': 450000.00,
            'dias_inicio': -365,
            'dias_fim': -30,
            'progresso': 100,
            'responsavel': 'Eng. Roberto Lima',
            'descricao': 'Galpão de 2000m² para armazenagem e distribuição',
        },
        {
            'nome': 'Edifício Corporate Tower',
            'cliente': 'Empreendimentos XYZ',
            'endereco': 'Av. Paulista, 2000 - São Paulo, SP',
            'status': 'Em Execução',
            'orcamento': 8000000.00,
            'dias_inicio': -240,
            'dias_fim': 240,
            'progresso': 50,
            'responsavel': 'Eng. Fernanda Souza',
            'descricao': 'Edifício comercial de 30 andares com escritórios',
        },
        {
            'nome': 'Residencial Parque Verde',
            'cliente': 'Construtora Horizonte',
            'endereco': 'Rua do Parque, 789 - Florianópolis, SC',
            'status': 'Paralisada',
            'orcamento': 600000.00,
            'dias_inicio': -90,
            'dias_fim': 120,
            'progresso': 30,
            'responsavel': 'Eng. Marcos Pereira',
            'descricao': 'Conjunto residencial com 8 casas geminadas',
        },
        {
            'nome': 'Shopping Center Metropolitano',
            'cliente': 'Grupo Shopping Brasil',
            'endereco': 'Av. das Américas, 3000 - Salvador, BA',
            'status': 'Planejamento',
            'orcamento': 15000000.00,
            'dias_inicio': 90,
            'dias_fim': 1095,
            'progresso': 0,
            'responsavel': 'Eng. Juliana Martins',
            'descricao': 'Shopping center com 150 lojas e 4 cinemas',
        },
        {
            'nome': 'Estação de Tratamento de Água',
            'cliente': 'Companhia de Saneamento',
            'endereco': 'Estrada da Água, S/N - Campinas, SP',
            'status': 'Em Execução',
            'orcamento': 3200000.00,
            'dias_inicio': -300,
            'dias_fim': 60,
            'progresso': 80,
            'responsavel': 'Eng. Ricardo Almeida',
            'descricao': 'ETA com capacidade de 500m³/hora',
        },
        {
            'nome': 'Hotel Resort Tropical',
            'cliente': 'Turismo e Lazer S/A',
            'endereco': 'Rodovia do Sol, Km 25 - Fortaleza, CE',
            'status': 'Entregue',
            'orcamento': 5500000.00,
            'dias_inicio': -540,
            'dias_fim': -60,
            'progresso': 100,
            'responsavel': 'Eng. Camila Ferreira',
            'descricao': 'Hotel resort com 200 quartos e área de lazer',
        },
        {
            'nome': 'Fábrica de Componentes Eletrônicos',
            'cliente': 'Tech Industry Ltda',
            'endereco': 'Distrito Industrial, Lote 45 - Manaus, AM',
            'status': 'Em Execução',
            'orcamento': 2800000.00,
            'dias_inicio': -150,
            'dias_fim': 210,
            'progresso': 45,
            'responsavel': 'Eng. Bruno Carvalho',
            'descricao': 'Unidade fabril de 5000m² com linha de montagem',
        },
        {
            'nome': 'Centro de Convenções Municipal',
            'cliente': 'Prefeitura de Goiânia',
            'endereco': 'Av. do Evento, 1500 - Goiânia, GO',
            'status': 'Planejamento',
            'orcamento': 4200000.00,
            'dias_inicio': 45,
            'dias_fim': 540,
            'progresso': 8,
            'responsavel': 'Eng. Patricia Lima',
            'descricao': 'Centro de convenções com capacidade para 5000 pessoas',
        },
        {
            'nome': 'Condomínio Horizontal Solar',
            'cliente': 'Incorporadora Sol Nascente',
            'endereco': 'Alameda das Palmeiras, 200 - Natal, RN',
            'status': 'Em Execução',
            'orcamento': 1800000.00,
            'dias_inicio': -210,
            'dias_fim': 150,
            'progresso': 55,
            'responsavel': 'Eng. Gustavo Henrique',
            'descricao': 'Condomínio com 20 casas e área comum',
        },
        {
            'nome': 'Terminal Rodoviário Urbano',
            'cliente': 'Secretaria de Transportes',
            'endereco': 'Praça da Mobilidade, S/N - Recife, PE',
            'status': 'Concluída',
            'orcamento': 950000.00,
            'dias_inicio': -400,
            'dias_fim': -45,
            'progresso': 100,
            'responsavel': 'Eng. Luciana Torres',
            'descricao': 'Terminal com 15 plataformas e estacionamento',
        },
        {
            'nome': 'Indústria de Alimentos Frescos',
            'cliente': 'Alimentos do Brasil S/A',
            'endereco': 'Rodovia dos Alimentos, Km 12 - Londrina, PR',
            'status': 'Em Execução',
            'orcamento': 3800000.00,
            'dias_inicio': -270,
            'dias_fim': 120,
            'progresso': 70,
            'responsavel': 'Eng. Daniel Moreira',
            'descricao': 'Unidade industrial de processamento de alimentos',
        },
        {
            'nome': 'Complexo Esportivo Municipal',
            'cliente': 'Secretaria de Esportes',
            'endereco': 'Rua do Esporte, 500 - São Luís, MA',
            'status': 'Paralisada',
            'orcamento': 2100000.00,
            'dias_inicio': -180,
            'dias_fim': 180,
            'progresso': 35,
            'responsavel': 'Eng. Sandra Melo',
            'descricao': 'Complexo com ginásio, piscina e campos',
        },
        {
            'nome': 'Torre de Escritórios Business',
            'cliente': 'Investimentos Imobiliários',
            'endereco': 'Av. Faria Lima, 3500 - São Paulo, SP',
            'status': 'Planejamento',
            'orcamento': 12000000.00,
            'dias_inicio': 120,
            'dias_fim': 900,
            'progresso': 2,
            'responsavel': 'Eng. Alexandre Costa',
            'descricao': 'Torre corporativa de 40 andares certificada LEED',
        },
        {
            'nome': 'Residencial Jardim das Acácias',
            'cliente': 'Construtora Familiar',
            'endereco': 'Rua das Acácias, 100 - Vitória, ES',
            'status': 'Em Execução',
            'orcamento': 720000.00,
            'dias_inicio': -100,
            'dias_fim': 200,
            'progresso': 35,
            'responsavel': 'Eng. Bianca Andrade',
            'descricao': 'Edifício residencial com 24 apartamentos',
        },
        {
            'nome': 'Aeroporto Regional Executivo',
            'cliente': 'Secretaria de Aviação',
            'endereco': 'Rodovia do Aeroporto, Km 8 - Ribeirão Preto, SP',
            'status': 'Entregue',
            'orcamento': 8500000.00,
            'dias_inicio': -730,
            'dias_fim': -90,
            'progresso': 100,
            'responsavel': 'Eng. Felipe Rocha',
            'descricao': 'Aeroporto com pista de 2000m e terminal',
        },
        {
            'nome': 'Usina Solar Fotovoltaica',
            'cliente': 'Energia Limpa Brasil',
            'endereco': 'Zona Rural, Gleba 45 - Juazeiro, BA',
            'status': 'Em Execução',
            'orcamento': 15000000.00,
            'dias_inicio': -365,
            'dias_fim': 365,
            'progresso': 40,
            'responsavel': 'Eng. Carolina Dias',
            'descricao': 'Usina solar com capacidade de 50MW',
        },
    ]

    # Preencher dados
    hoje = date.today()

    for row_idx, obra in enumerate(obras, 2):
        data_inicio = hoje + timedelta(days=obra['dias_inicio'])
        data_fim = hoje + timedelta(days=obra['dias_fim'])

        ws.cell(row=row_idx, column=1, value=obra['nome'])
        ws.cell(row=row_idx, column=2, value=obra['cliente'])
        ws.cell(row=row_idx, column=3, value=obra['endereco'])
        ws.cell(row=row_idx, column=4, value=obra['status'])
        ws.cell(row=row_idx, column=5, value=obra['orcamento'])
        ws.cell(row=row_idx, column=6, value=data_inicio)
        ws.cell(row=row_idx, column=7, value=data_fim)
        ws.cell(row=row_idx, column=8, value=obra['progresso'])
        ws.cell(row=row_idx, column=9, value=obra['responsavel'])
        ws.cell(row=row_idx, column=10, value=obra['descricao'])

        # Aplicar bordas
        for col_idx in range(1, 11):
            ws.cell(row=row_idx, column=col_idx).border = border

    # Formatar coluna de data
    for row in range(2, len(obras) + 2):
        ws.cell(row=row, column=6).number_format = 'DD/MM/YYYY'
        ws.cell(row=row, column=7).number_format = 'DD/MM/YYYY'
        ws.cell(row=row, column=5).number_format = '#,##0.00'

    # Ajustar largura das colunas
    larguras = [30, 25, 40, 15, 20, 15, 15, 12, 25, 40]
    for idx, largura in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(idx)].width = largura

    # Congelar primeira linha
    ws.freeze_panes = 'A2'

    # Auto-filtro
    ws.auto_filter.ref = f'A1:J{len(obras) + 1}'

    # Salvar em memória
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def gerar_planilha_lancamentos_exemplo():
    """
    Gera uma planilha Excel de exemplo com lançamentos financeiros
    para as 20 obras
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Lançamentos Exemplo'

    # Estilos
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    # Cabeçalhos
    cabecalhos = [
        'Obra',
        'Descrição',
        'Categoria',
        'Tipo',
        'Valor (R$)',
        'Data',
        'Status Pagamento',
    ]
    for col_idx, titulo in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Dados de exemplo - lançamentos para as obras
    categorias_despesa = [
        'Material',
        'Mão de Obra',
        'Equipamento',
        'Serviço Terceiro',
        'Imposto',
        'Outros',
    ]
    categorias_receita = ['Venda', 'Prestação de Serviço', 'Outros']
    status_list = ['Pago', 'Pendente', 'Pago']

    obras_nomes = [
        'Residencial Villa Nova',
        'Centro Comercial Plaza',
        'Escola Municipal Prof. Lima',
        'Hospital Regional Norte',
        'Galpão Industrial Logística',
        'Edifício Corporate Tower',
        'Residencial Parque Verde',
        'Shopping Center Metropolitano',
        'Estação de Tratamento de Água',
        'Hotel Resort Tropical',
        'Fábrica de Componentes Eletrônicos',
        'Centro de Convenções Municipal',
        'Condomínio Horizontal Solar',
        'Terminal Rodoviário Urbano',
        'Indústria de Alimentos Frescos',
        'Complexo Esportivo Municipal',
        'Torre de Escritórios Business',
        'Residencial Jardim das Acácias',
        'Aeroporto Regional Executivo',
        'Usina Solar Fotovoltaica',
    ]

    hoje = date.today()
    row = 2

    # Gerar lançamentos para cada obra
    for obra_nome in obras_nomes:
        # 3-5 lançamentos de despesa por obra
        num_despesas = random.randint(3, 5)
        for _ in range(num_despesas):
            ws.cell(row=row, column=1, value=obra_nome)
            ws.cell(
                row=row,
                column=2,
                value=f'{random.choice(["Compra", "Pagamento", "Serviço"])} - {random.choice(["fase 1", "fase 2", "parcial"])}',
            )
            ws.cell(row=row, column=3, value=random.choice(categorias_despesa))
            ws.cell(row=row, column=4, value='Despesa')
            ws.cell(row=row, column=5, value=round(random.uniform(5000, 150000), 2))
            ws.cell(row=row, column=6, value=hoje - timedelta(days=random.randint(1, 180)))
            ws.cell(row=row, column=7, value=random.choice(status_list))
            row += 1

        # 1-2 lançamentos de receita por obra
        num_receitas = random.randint(1, 2)
        for _ in range(num_receitas):
            ws.cell(row=row, column=1, value=obra_nome)
            ws.cell(
                row=row,
                column=2,
                value=f'{random.choice(["Recebimento", "Entrada", "Pagamento cliente"])}',
            )
            ws.cell(row=row, column=3, value=random.choice(categorias_receita))
            ws.cell(row=row, column=4, value='Receita')
            ws.cell(row=row, column=5, value=round(random.uniform(10000, 300000), 2))
            ws.cell(row=row, column=6, value=hoje - timedelta(days=random.randint(1, 180)))
            ws.cell(row=row, column=7, value='Pago')
            row += 1

    # Formatar
    for r in range(2, row):
        ws.cell(row=r, column=5).number_format = '#,##0.00'
        ws.cell(row=r, column=6).number_format = 'DD/MM/YYYY'

    # Ajustar largura
    larguras = [30, 30, 20, 12, 15, 15, 18]
    for idx, largura in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(idx)].width = largura

    ws.freeze_panes = 'A2'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
