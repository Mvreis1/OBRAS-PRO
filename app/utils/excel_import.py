"""
Helper para importacao de dados de Excel (.xlsx, .csv)
"""
import csv
import io
from datetime import datetime, date
from openpyxl import load_workbook


class ExcelImportError(Exception):
    """Excecao para erros de importacao"""
    pass


def parse_date(value):
    """Tenta converter valor para data em varios formatos"""
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    if not value or str(value).strip() == '':
        return None

    formats = [
        '%d/%m/%Y',
        '%Y-%m-%d',
        '%d-%m-%Y',
        '%m/%d/%Y',
        '%d/%m/%y',
        '%Y/%m/%d',
    ]

    value_str = str(value).strip()
    for fmt in formats:
        try:
            return datetime.strptime(value_str, fmt).date()
        except ValueError:
            continue

    # Tenta converter de numero Excel (dias desde 1899-12-30)
    try:
        from datetime import timedelta
        excel_base = date(1899, 12, 30)
        days = int(float(value_str))
        return excel_base + timedelta(days=days)
    except (ValueError, TypeError):
        pass

    return None


def parse_float(value):
    """Converte valor para float, tratando formatos brasileiros"""
    if isinstance(value, (int, float)):
        return float(value)
    if not value or str(value).strip() == '':
        return 0.0

    value_str = str(value).strip()

    # Remove simbolo de moeda e espacos
    value_str = value_str.replace('R$', '').replace('$', '').strip()

    # Detecta formato: se tem ponto e virgula, provavelmente eh brasileiro
    if ',' in value_str and '.' in value_str:
        # Formato brasileiro: 1.234,56 ou americano: 1,234.56
        last_comma = value_str.rfind(',')
        last_dot = value_str.rfind('.')
        if last_comma > last_dot:
            # Brasileiro: 1.234,56
            value_str = value_str.replace('.', '').replace(',', '.')
        else:
            # Americano: 1,234.56
            value_str = value_str.replace(',', '')
    elif ',' in value_str:
        # So tem virgula - pode ser decimal brasileiro
        value_str = value_str.replace(',', '.')

    try:
        return float(value_str)
    except ValueError:
        return 0.0


def importar_lancamentos_excel(file_stream, filename):
    """
    Importa lancamentos de arquivo Excel ou CSV.

    Retorna:
        tuple: (lancamentos_list, erros_list)
        lancamentos_list: lista de dicts com os dados
        erros_list: lista de mensagens de erro por linha
    """
    lancamentos = []
    erros = []

    if filename.lower().endswith('.csv'):
        return _importar_csv(file_stream)
    elif filename.lower().endswith('.xlsx'):
        return _importar_xlsx(file_stream)
    else:
        raise ExcelImportError("Formato nao suportado. Use .xlsx ou .csv")


def _importar_csv(file_stream):
    """Importa dados de CSV"""
    lancamentos = []
    erros = []

    # Detecta encoding
    content = file_stream.read()
    try:
        text = content.decode('utf-8')
    except UnicodeDecodeError:
        try:
            text = content.decode('latin-1')
        except UnicodeDecodeError:
            raise ExcelImportError("Nao foi possivel decodificar o arquivo. Use UTF-8 ou Latin-1.")

    # Detecta delimitador
    sample = text[:1024]
    delimiter = ';' if sample.count(';') > sample.count(',') else ','

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)

    for idx, row in enumerate(reader, start=2):  # Linha 1 = header
        try:
            lancamento = _parse_row(row, idx)
            if lancamento:
                lancamentos.append(lancamento)
        except Exception as e:
            erros.append(f"Linha {idx}: {str(e)}")

    return lancamentos, erros


def _importar_xlsx(file_stream):
    """Importa dados de Excel (.xlsx)"""
    lancamentos = []
    erros = []

    try:
        wb = load_workbook(file_stream, data_only=True)
        ws = wb.active

        # Detecta cabecalhos
        headers = {}
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        header_map = {
            'obra': ['obra', 'nome obra', 'obra nome', 'projeto', 'nome'],
            'descricao': ['descricao', 'desc', 'historico', 'detalhe', 'observacao'],
            'categoria': ['categoria', 'tipo', 'classificacao', 'grupo'],
            'tipo': ['tipo', 'receita/despesa', 'natureza', 'tipo lancamento'],
            'valor': ['valor', 'valor r$', 'valor rs', 'montante', 'quantia'],
            'data': ['data', 'data lancamento', 'data pagamento', 'vencimento', 'date'],
            'forma_pagamento': ['forma pagamento', 'pagamento', 'metodo', 'forma'],
            'status_pagamento': ['status', 'status pagamento', 'situacao', 'pago'],
            'documento': ['documento', 'doc', 'numero', 'numero doc', 'referencia'],
            'observacoes': ['observacoes', 'obs', 'notas', 'comentarios'],
        }

        for col_idx, header in enumerate(header_row, start=1):
            if not header:
                continue
            header_lower = str(header).lower().strip()
            for key, variants in header_map.items():
                if header_lower in variants:
                    headers[key] = col_idx - 1  # indice 0-based
                    break

        # Verifica campos obrigatorios
        required = ['descricao', 'valor', 'data']
        missing = [r for r in required if r not in headers]
        if missing:
            raise ExcelImportError(f"Colunas obrigatorias nao encontradas: {', '.join(missing)}")

        # Processa linhas
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Pula linhas vazias
                if not row or all(v is None or str(v).strip() == '' for v in row):
                    continue

                lancamento = _parse_row_from_indices(row, headers, row_idx)
                if lancamento:
                    lancamentos.append(lancamento)
            except Exception as e:
                erros.append(f"Linha {row_idx}: {str(e)}")

    except Exception as e:
        raise ExcelImportError(f"Erro ao ler arquivo Excel: {str(e)}")

    return lancamentos, erros


def _parse_row(row_dict, line_num):
    """Parse de uma linha do CSV (dict)"""
    # Normaliza keys
    normalized = {}
    for key, value in row_dict.items():
        if key:
            normalized[key.lower().strip()] = value

    # Campos obrigatorios
    descricao = normalized.get('descricao') or normalized.get('desc') or normalized.get('historico')
    if not descricao or str(descricao).strip() == '':
        raise ValueError("Descricao e obrigatoria")

    valor_str = normalized.get('valor') or normalized.get('valor r$')
    valor = parse_float(valor_str)
    if valor == 0:
        raise ValueError("Valor invalido ou zero")

    data_str = normalized.get('data') or normalized.get('data lancamento') or normalized.get('vencimento')
    data = parse_date(data_str)
    if not data:
        raise ValueError(f"Data invalida: {data_str}")

    # Detecta tipo (Receita/Despesa)
    tipo = normalized.get('tipo', '').strip()
    if tipo.lower() in ['receita', 'receitas', 'entrada', 'entradas', 'credito', 'creditos']:
        tipo = 'Receita'
    elif tipo.lower() in ['despesa', 'despesas', 'saida', 'saidas', 'debito', 'debitos']:
        tipo = 'Despesa'
    else:
        # Tenta inferir pelo valor
        tipo = 'Receita' if valor > 0 else 'Despesa'

    valor = abs(valor)  # Sempre positivo no banco

    return {
        'obra_nome': normalized.get('obra', ''),
        'descricao': str(descricao).strip(),
        'categoria': normalized.get('categoria', 'Geral'),
        'tipo': tipo,
        'valor': valor,
        'data': data,
        'forma_pagamento': normalized.get('forma_pagamento') or normalized.get('forma pagamento', 'Transferencia'),
        'status_pagamento': normalized.get('status_pagamento') or normalized.get('status', 'Pago'),
        'documento': normalized.get('documento', ''),
        'observacoes': normalized.get('observacoes') or normalized.get('obs', ''),
    }


def _parse_row_from_indices(row_tuple, headers, line_num):
    """Parse de uma linha do Excel (tuple) usando indices"""
    def get_val(key, default=''):
        if key not in headers:
            return default
        idx = headers[key]
        if idx < len(row_tuple):
            val = row_tuple[idx]
            return val if val is not None else default
        return default

    # Campos obrigatorios
    descricao = get_val('descricao')
    if not descricao or str(descricao).strip() == '':
        raise ValueError("Descricao e obrigatoria")

    valor = parse_float(get_val('valor', 0))
    if valor == 0:
        raise ValueError("Valor invalido ou zero")

    data = parse_date(get_val('data'))
    if not data:
        raise ValueError("Data invalida")

    # Detecta tipo
    tipo_val = str(get_val('tipo', '')).strip()
    if tipo_val.lower() in ['receita', 'receitas', 'entrada', 'entradas', 'credito']:
        tipo = 'Receita'
    elif tipo_val.lower() in ['despesa', 'despesas', 'saida', 'saidas', 'debito']:
        tipo = 'Despesa'
    else:
        tipo = 'Receita' if valor > 0 else 'Despesa'

    valor = abs(valor)

    return {
        'obra_nome': str(get_val('obra', '')).strip(),
        'descricao': str(descricao).strip(),
        'categoria': str(get_val('categoria', 'Geral')).strip(),
        'tipo': tipo,
        'valor': valor,
        'data': data,
        'forma_pagamento': str(get_val('forma_pagamento', 'Transferencia')).strip(),
        'status_pagamento': str(get_val('status_pagamento', 'Pago')).strip(),
        'documento': str(get_val('documento', '')).strip(),
        'observacoes': str(get_val('observacoes', '')).strip(),
    }


def gerar_modelo_excel():
    """Gera arquivo Excel modelo para importacao"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Lancamentos"

    # Cabecalhos
    headers = ['Obra', 'Descricao', 'Categoria', 'Tipo', 'Valor', 'Data', 'Forma Pagamento', 'Status', 'Documento', 'Observacoes']
    ws.append(headers)

    # Estilo do cabecalho
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Dados de exemplo
    exemplo1 = ['Edificio Alpha', 'Compra de cimento', 'Materiais', 'Despesa', 5000.00, '15/01/2026', 'Transferencia', 'Pago', 'NF-123', 'Pagamento a vista']
    exemplo2 = ['Edificio Alpha', 'Receita de vendas', 'Vendas', 'Receita', 15000.00, '20/01/2026', 'Boleto', 'Pago', '', 'Parcela 1']
    exemplo3 = ['Residencial Parque', 'Mao de obra', 'Servicos', 'Despesa', 3500.00, '10/01/2026', 'PIX', 'Pago', '', 'Pedreiro']

    ws.append(exemplo1)
    ws.append(exemplo2)
    ws.append(exemplo3)

    # Ajusta largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Congela primeira linha
    ws.freeze_panes = 'A2'

    # Salva em memoria
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def importar_obras_excel(file_stream, filename):
    """
    Importa obras de arquivo Excel.

    Retorna:
        tuple: (obras_list, erros_list)
    """
    obras = []
    erros = []

    if filename.lower().endswith('.xlsx'):
        return _importar_obras_xlsx(file_stream)
    else:
        raise ExcelImportError("Formato nao suportado para obras. Use .xlsx")


def _importar_obras_xlsx(file_stream):
    """Importa obras de Excel (.xlsx)"""
    obras = []
    erros = []

    try:
        wb = load_workbook(file_stream, data_only=True)
        ws = wb.active

        # Detecta cabecalhos
        headers = {}
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        header_map = {
            'nome': ['nome', 'nome da obra', 'obra', 'nome obra', 'titulo'],
            'cliente': ['cliente', 'nome cliente', 'cliente nome', 'proprietario'],
            'endereco': ['endereco', 'end', 'local', 'localizacao', 'rua'],
            'status': ['status', 'situacao', 'estado', 'fase'],
            'orcamento_previsto': ['orcamento', 'orcamento previsto', 'valor', 'valor previsto', 'custo'],
            'data_inicio': ['data inicio', 'inicio', 'data de inicio', 'start'],
            'data_fim_prevista': ['data fim', 'fim', 'data fim prevista', 'previsao', 'data prevista'],
            'progresso': ['progresso', 'percentual', 'andamento', '%'],
            'responsavel': ['responsavel', 'engenheiro', 'gestor', 'responsavel tecnico'],
            'descricao': ['descricao', 'desc', 'detalhes', 'observacoes'],
        }

        for col_idx, header in enumerate(header_row, start=1):
            if not header:
                continue
            header_lower = str(header).lower().strip()
            for key, variants in header_map.items():
                if header_lower in variants:
                    headers[key] = col_idx - 1
                    break

        # Verifica campos obrigatorios
        required = ['nome']
        missing = [r for r in required if r not in headers]
        if missing:
            raise ExcelImportError(f"Colunas obrigatorias nao encontradas: {', '.join(missing)}")

        # Processa linhas
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Pula linhas vazias
                if not row or all(v is None or str(v).strip() == '' for v in row):
                    continue

                obra = _parse_obra_row(row, headers, row_idx)
                if obra:
                    obras.append(obra)
            except Exception as e:
                erros.append(f"Linha {row_idx}: {str(e)}")

    except Exception as e:
        raise ExcelImportError(f"Erro ao ler arquivo Excel: {str(e)}")

    return obras, erros


def _parse_obra_row(row_tuple, headers, line_num):
    """Parse de uma linha de obra do Excel"""
    def get_val(key, default=''):
        if key not in headers:
            return default
        idx = headers[key]
        if idx < len(row_tuple):
            val = row_tuple[idx]
            return val if val is not None else default
        return default

    # Campo obrigatorio
    nome = get_val('nome')
    if not nome or str(nome).strip() == '':
        raise ValueError("Nome da obra e obrigatorio")

    # Valida status
    status = str(get_val('status', 'Planejamento')).strip()
    status_validos = ['Planejamento', 'Em Execução', 'Paralisada', 'Concluída', 'Entregue']
    if status not in status_validos:
        status = 'Planejamento'

    # Parse valores
    orcamento = parse_float(get_val('orcamento_previsto', 0))
    progresso = int(parse_float(get_val('progresso', 0)))
    if progresso < 0:
        progresso = 0
    if progresso > 100:
        progresso = 100

    # Parse datas
    data_inicio = parse_date(get_val('data_inicio'))
    data_fim = parse_date(get_val('data_fim_prevista'))

    return {
        'nome': str(nome).strip(),
        'cliente': str(get_val('cliente', '')).strip(),
        'endereco': str(get_val('endereco', '')).strip(),
        'status': status,
        'orcamento_previsto': orcamento,
        'data_inicio': data_inicio,
        'data_fim_prevista': data_fim,
        'progresso': progresso,
        'responsavel': str(get_val('responsavel', '')).strip(),
        'descricao': str(get_val('descricao', '')).strip(),
    }
